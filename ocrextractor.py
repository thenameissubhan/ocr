"""
PHASE 2: ROW OCR EXTRACTOR  +  GEMINI AI CORRECTION  (v4)
==========================================================
EasyOCR extracts rows  ->  Gemini Flash sees the image and fixes OCR errors.

Requirements:
    pip install easyocr opencv-python pillow pandas openpyxl numpy google-generativeai
"""

import warnings
warnings.filterwarnings("ignore", message=".*pin_memory.*")
warnings.filterwarnings("ignore", category=UserWarning)

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import cv2
import os
import sys
import subprocess
import threading
import base64
import json
import re
import numpy as np
import pandas as pd
from pathlib import Path
from PIL import Image, ImageTk
import easyocr

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Gemini
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False


# ── colours ──────────────────────────────────────────────
C_HEADER = "#1a1d2e"
C_BG     = "#f0f2f5"
C_ACCENT = "#00b894"
C_BLUE   = "#0984e3"
C_ORANGE = "#e17055"
C_PURPLE = "#6c5ce7"
C_RED    = "#d63031"
C_GOLD   = "#f9ca24"
C_AI     = "#a29bfe"        # colour for AI-corrected rows
C_TEXT   = "#2d3436"
C_GRAY   = "#636e72"
C_BORDER = "#dce1e7"

ROW_COLORS = [
    (0, 184, 148), (255, 107, 53), (108, 92, 231),
    (253, 203, 110), (9, 132, 227), (225, 112, 85),
]
ROW_AI_COLOR = (162, 155, 254)   # violet for AI-fixed rows


def make_scrollable_frame(parent):
    outer  = tk.Frame(parent, bg=C_BG)
    canvas = tk.Canvas(outer, bg=C_BG, highlightthickness=0, bd=0)
    sb     = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
    canvas.configure(yscrollcommand=sb.set)
    sb.pack(side=tk.RIGHT, fill=tk.Y)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    inner     = tk.Frame(canvas, bg=C_BG)
    window_id = canvas.create_window((0, 0), window=inner, anchor="nw")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(window_id, width=e.width))
    def _scroll(ev):
        if ev.delta:        canvas.yview_scroll(int(-ev.delta / 120), "units")
        elif ev.num == 4:   canvas.yview_scroll(-1, "units")
        elif ev.num == 5:   canvas.yview_scroll(1,  "units")
    canvas.bind_all("<MouseWheel>", _scroll)
    canvas.bind_all("<Button-4>",   _scroll)
    canvas.bind_all("<Button-5>",   _scroll)
    return outer, inner


# ═══════════════════════════════════════════════════════════════════════
class Phase2OCRExtractor:
    def __init__(self, root):
        self.root = root
        self.root.title("Phase 2: Row OCR + Gemini AI Correction  ->  Excel")
        self.root.minsize(1050, 660)
        try:    self.root.state("zoomed")
        except: self.root.attributes("-zoomed", True)

        # OCR settings
        self.input_folder    = tk.StringVar()
        self.output_folder   = tk.StringVar()
        self.row_tolerance   = tk.IntVar(value=14)
        self.min_confidence  = tk.DoubleVar(value=0.25)
        self.lang_choice     = tk.StringVar(value="en")
        self.gpu_enabled     = tk.BooleanVar(value=False)
        self.show_boxes      = tk.BooleanVar(value=True)
        self.show_text_ov    = tk.BooleanVar(value=True)
        self.text_threshold  = tk.DoubleVar(value=0.4)
        self.width_threshold = tk.DoubleVar(value=0.5)
        self.decoder_choice  = tk.StringVar(value="greedy")
        self.min_row_len     = tk.IntVar(value=5)

        # AI settings
        self.ai_enabled      = tk.BooleanVar(value=False)
        self.gemini_api_key  = tk.StringVar()
        self.gemini_model    = tk.StringVar(value="gemini-2.0-flash")
        self._gemini_client  = None

        # ROI
        self._roi         = None
        self._roi_drawing = False
        self._roi_start   = None
        self._roi_rect_id = None
        self._disp_scale  = 1.0
        self._disp_offset = (0, 0)

        self.image_files        = []
        self.current_idx        = 0
        self.all_results        = []
        self.reader             = None
        self.processing         = False
        self._current_cv_img    = None
        self._current_row_data  = []
        self._tk_img            = None
        self._last_excel_path   = None
        self._show_ai_diff      = tk.BooleanVar(value=True)

        self._build_ui()

    # ═══════════════════════ UI ═══════════════════════

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg=C_HEADER)
        hdr.pack(fill=tk.X, side=tk.TOP)
        tk.Label(hdr, text="PHASE 2  \u2014  OCR  +  AI CORRECTION  \u2192  EXCEL",
                 font=("Segoe UI", 14, "bold"),
                 bg=C_HEADER, fg=C_ACCENT, pady=10, padx=14).pack(side=tk.LEFT)
        tk.Label(hdr, text="EasyOCR  |  Gemini Flash  |  Row Detection  |  Excel",
                 font=("Segoe UI", 9), bg=C_HEADER, fg="#b2bec3").pack(side=tk.LEFT)

        # Status bar
        bot = tk.Frame(self.root, bg=C_HEADER, pady=3)
        bot.pack(fill=tk.X, side=tk.BOTTOM)
        self.progress_bar = ttk.Progressbar(bot, orient=tk.HORIZONTAL,
                                             mode="determinate", length=300)
        self.progress_bar.pack(side=tk.LEFT, padx=8, pady=3)
        self.lbl_status = tk.Label(bot, text="Ready",
                                    bg=C_HEADER, fg=C_ACCENT,
                                    font=("Segoe UI", 9), padx=8)
        self.lbl_status.pack(side=tk.LEFT)
        self.lbl_row_count = tk.Label(bot, text="",
                                       bg=C_HEADER, fg="#b2bec3",
                                       font=("Segoe UI", 9))
        self.lbl_row_count.pack(side=tk.RIGHT, padx=8)
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TProgressbar", troughcolor="#2d3561",
                        background=C_ACCENT, thickness=14)

        # Body
        body = tk.Frame(self.root, bg=C_BG)
        body.pack(fill=tk.BOTH, expand=True)

        scroll_outer, left = make_scrollable_frame(body)
        scroll_outer.configure(width=345)
        scroll_outer.pack(side=tk.LEFT, fill=tk.Y)
        scroll_outer.pack_propagate(False)

        right = tk.Frame(body, bg="#12122a")
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Nav
        nav = tk.Frame(right, bg="#1e1e40", pady=5, padx=10)
        nav.pack(fill=tk.X)
        self._nbtn(nav, "PREV", self.prev_image).pack(side=tk.LEFT, padx=(0, 4))
        self._nbtn(nav, "NEXT", self.next_image).pack(side=tk.LEFT, padx=(0, 10))
        self.lbl_filename = tk.Label(nav, text="No image loaded",
                                      bg="#1e1e40", fg=C_ACCENT,
                                      font=("Segoe UI", 10, "bold"))
        self.lbl_filename.pack(side=tk.LEFT)
        self.lbl_progress_nav = tk.Label(nav, text="", bg="#1e1e40",
                                          fg="#b2bec3", font=("Segoe UI", 9))
        self.lbl_progress_nav.pack(side=tk.RIGHT)

        # Canvas
        self.canvas = tk.Canvas(right, bg="#0d0d1a", highlightthickness=0,
                                 cursor="crosshair")
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=6, pady=(4, 0))
        self.canvas.bind("<ButtonPress-1>",   self._roi_on_press)
        self.canvas.bind("<B1-Motion>",       self._roi_on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._roi_on_release)

        self.lbl_roi_info = tk.Label(right,
                                      text="  Drag on image to select data region (ROI)",
                                      bg="#1e1e40", fg="#fdcb6e",
                                      font=("Segoe UI", 8), anchor="w", pady=2)
        self.lbl_roi_info.pack(fill=tk.X, padx=6)

        # ── TABBED TEXT OUTPUT  (OCR raw  |  AI corrected) ──
        tab_bar = tk.Frame(right, bg="#1e1e40")
        tab_bar.pack(fill=tk.X, padx=6, pady=(2, 0))

        self._active_tab = tk.StringVar(value="ocr")

        def tab_btn(text, val):
            def _sel():
                self._active_tab.set(val)
                _refresh_tabs()
            b = tk.Button(tab_bar, text=text,
                           font=("Segoe UI", 8, "bold"),
                           relief=tk.FLAT, bd=0, padx=12, pady=4,
                           cursor="hand2", command=_sel)
            b.pack(side=tk.LEFT)
            return b

        self.tab_ocr = tab_btn("RAW OCR", "ocr")
        self.tab_ai  = tab_btn("AI CORRECTED", "ai")

        def _refresh_tabs():
            v = self._active_tab.get()
            self.tab_ocr.config(bg=C_BLUE   if v == "ocr" else "#2d2d50",
                                 fg="white")
            self.tab_ai.config( bg=C_AI     if v == "ai"  else "#2d2d50",
                                fg="#1a1a2e" if v == "ai" else "#b2bec3")
            if v == "ocr":
                self.frame_txt_ai.pack_forget()
                self.frame_txt_ocr.pack(fill=tk.BOTH, expand=True)
            else:
                self.frame_txt_ocr.pack_forget()
                self.frame_txt_ai.pack(fill=tk.BOTH, expand=True)

        self._refresh_tabs = _refresh_tabs

        txt_container = tk.Frame(right, bg="#1e1e40", height=165)
        txt_container.pack(fill=tk.X, padx=6, pady=(0, 4))
        txt_container.pack_propagate(False)

        def make_txt(parent, bg_color):
            f = tk.Frame(parent, bg=bg_color)
            scr = tk.Scrollbar(f)
            scr.pack(side=tk.RIGHT, fill=tk.Y)
            t = tk.Text(f, bg="#0d0d1a", fg="#dfe6e9",
                        font=("Consolas", 9),
                        yscrollcommand=scr.set,
                        state="disabled", padx=8, pady=4)
            t.pack(fill=tk.BOTH, expand=True)
            scr.config(command=t.yview)
            return f, t

        self.frame_txt_ocr, self.txt_ocr = make_txt(txt_container, "#1e1e40")
        self.frame_txt_ai,  self.txt_ai  = make_txt(txt_container, "#1e1e40")
        _refresh_tabs()

        # ══════════ LEFT PANEL ══════════

        def section(title, color=C_BLUE):
            f = tk.LabelFrame(left, text=f"  {title}  ",
                               font=("Segoe UI", 9, "bold"),
                               fg=color, bg=C_BG, relief=tk.GROOVE,
                               bd=1, padx=10, pady=8)
            f.pack(fill=tk.X, padx=10, pady=(8, 0))
            return f

        def btn(parent, text, color, cmd, py=7):
            return tk.Button(parent, text=text,
                              font=("Segoe UI", 9, "bold"),
                              bg=color, fg="white",
                              activebackground="#111", activeforeground="white",
                              relief=tk.FLAT, bd=0, pady=py,
                              cursor="hand2", command=cmd)

        def entry_browse(parent, var, cmd, label):
            tk.Label(parent, text=label, bg=C_BG, fg=C_GRAY,
                     font=("Segoe UI", 8)).pack(anchor="w")
            row = tk.Frame(parent, bg=C_BG)
            row.pack(fill=tk.X, pady=(2, 4))
            tk.Entry(row, textvariable=var, font=("Segoe UI", 9),
                     relief=tk.SOLID, bd=1, bg="white",
                     fg=C_TEXT).pack(side=tk.LEFT, fill=tk.X, expand=True)
            btn(row, "...", C_BLUE, cmd, py=2).pack(side=tk.RIGHT, padx=(4, 0))

        def slider(parent, label, var, frm, to_, res, note=""):
            lf = tk.Frame(parent, bg=C_BG)
            lf.pack(fill=tk.X, pady=(3, 0))
            hf = tk.Frame(lf, bg=C_BG)
            hf.pack(fill=tk.X)
            tk.Label(hf, text=label, bg=C_BG, fg=C_TEXT,
                     font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT)
            tk.Label(hf, textvariable=var, bg=C_BG, fg=C_BLUE,
                     font=("Segoe UI", 8, "bold"), width=5).pack(side=tk.RIGHT)
            if note:
                tk.Label(lf, text=note, bg=C_BG, fg=C_GRAY,
                         font=("Segoe UI", 7)).pack(anchor="w")
            tk.Scale(lf, variable=var, from_=frm, to=to_, resolution=res,
                     orient=tk.HORIZONTAL, bg=C_BG, fg=C_TEXT,
                     troughcolor=C_BORDER, highlightthickness=0,
                     showvalue=False,
                     activebackground=C_BLUE).pack(fill=tk.X)

        # 01 Input
        s1 = section("01  Input Folder", C_BLUE)
        entry_browse(s1, self.input_folder, self.browse_input, "Screenshots folder:")
        self.lbl_stats = tk.Label(s1, text="No folder loaded",
                                   bg=C_BG, fg=C_GRAY, font=("Segoe UI", 8),
                                   wraplength=280, justify=tk.LEFT)
        self.lbl_stats.pack(anchor="w")

        # 02 Output
        s2 = section("02  Output Folder", C_BLUE)
        entry_browse(s2, self.output_folder, self.browse_output, "Save Excel to:")
        self.lbl_out_path = tk.Label(s2, text="", bg=C_BG, fg=C_GRAY,
                                      font=("Segoe UI", 7),
                                      wraplength=280, justify=tk.LEFT)
        self.lbl_out_path.pack(anchor="w")

        # 03 ROI
        s3 = section("03  Data Region  (ROI)", "#fdcb6e")
        tk.Label(s3, text="Draw a box on the image to OCR ONLY\nthe data table area.",
                 bg=C_BG, fg=C_TEXT, font=("Segoe UI", 8),
                 justify=tk.LEFT).pack(anchor="w", pady=(0, 5))
        self.lbl_roi_coords = tk.Label(s3, text="No region set  (full image)",
                                        bg=C_BG, fg=C_GRAY, font=("Consolas", 8))
        self.lbl_roi_coords.pack(anchor="w")
        rb = tk.Frame(s3, bg=C_BG)
        rb.pack(fill=tk.X, pady=(5, 0))
        btn(rb, "CLEAR REGION", C_RED, self.clear_roi, py=4).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))

        # 04 AI CORRECTION  ← NEW SECTION
        s_ai = section("04  AI Correction  (Gemini)", C_AI)

        if not GEMINI_AVAILABLE:
            tk.Label(s_ai,
                     text="google-generativeai not installed.\nRun:  pip install google-generativeai",
                     bg=C_BG, fg=C_RED, font=("Segoe UI", 8),
                     justify=tk.LEFT).pack(anchor="w")
        
        # Enable toggle
        ai_toggle = tk.Frame(s_ai, bg=C_BG)
        ai_toggle.pack(fill=tk.X, pady=(0, 6))
        tk.Checkbutton(ai_toggle, text="Enable AI Correction",
                       variable=self.ai_enabled,
                       bg=C_BG, fg=C_TEXT, selectcolor="white",
                       font=("Segoe UI", 9, "bold"),
                       activebackground=C_BG,
                       command=self._toggle_ai_ui).pack(side=tk.LEFT)

        # AI sub-frame (hidden when disabled)
        self.ai_frame = tk.Frame(s_ai, bg=C_BG)
        self.ai_frame.pack(fill=tk.X)

        tk.Label(self.ai_frame,
                 text="Get free key:  aistudio.google.com",
                 bg=C_BG, fg=C_GRAY, font=("Segoe UI", 7)).pack(anchor="w")
        tk.Label(self.ai_frame, text="Gemini API Key:",
                 bg=C_BG, fg=C_TEXT, font=("Segoe UI", 8, "bold")).pack(anchor="w")
        key_row = tk.Frame(self.ai_frame, bg=C_BG)
        key_row.pack(fill=tk.X, pady=(2, 4))
        self.entry_key = tk.Entry(key_row, textvariable=self.gemini_api_key,
                                   font=("Consolas", 8),
                                   relief=tk.SOLID, bd=1,
                                   show="*",  # hide key
                                   bg="white", fg=C_TEXT)
        self.entry_key.pack(side=tk.LEFT, fill=tk.X, expand=True)
        btn(key_row, "Test", "#00cec9", self.test_ai_connection, py=2).pack(
            side=tk.RIGHT, padx=(4, 0))

        tk.Label(self.ai_frame, text="Model:", bg=C_BG, fg=C_TEXT,
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        model_f = tk.Frame(self.ai_frame, bg=C_BG)
        model_f.pack(fill=tk.X, pady=(2, 4))
        for m in ["gemini-2.0-flash", "gemini-1.5-flash", "gemini-2.0-flash-lite"]:
            tk.Radiobutton(model_f, text=m, variable=self.gemini_model, value=m,
                           bg=C_BG, fg=C_TEXT, selectcolor="white",
                           font=("Segoe UI", 7),
                           activebackground=C_BG).pack(anchor="w")

        tk.Checkbutton(self.ai_frame,
                       text="Highlight AI-changed cells in Excel",
                       variable=self._show_ai_diff,
                       bg=C_BG, fg=C_TEXT, selectcolor="white",
                       font=("Segoe UI", 8),
                       activebackground=C_BG).pack(anchor="w")

        self.lbl_ai_status = tk.Label(self.ai_frame, text="",
                                       bg=C_BG, fg=C_GRAY,
                                       font=("Segoe UI", 8),
                                       wraplength=280, justify=tk.LEFT)
        self.lbl_ai_status.pack(anchor="w", pady=(3, 0))

        self._toggle_ai_ui()   # start hidden

        # 05 OCR Settings
        s4 = section("05  OCR Settings", C_ORANGE)
        slider(s4, "Row merge tolerance (px):", self.row_tolerance, 4, 50, 1,
               "Increase if rows split apart")
        slider(s4, "Min confidence:", self.min_confidence, 0.05, 1.0, 0.05,
               "Lower = detect more text")
        slider(s4, "Text detection threshold:", self.text_threshold, 0.1, 0.9, 0.05,
               "Lower = catch faint text")
        slider(s4, "H-box merge threshold:", self.width_threshold, 0.1, 1.0, 0.05,
               "Merge nearby words on same line")
        slider(s4, "Min row length (chars):", self.min_row_len, 1, 30, 1,
               "Filter very short rows")

        tk.Label(s4, text="Language:", bg=C_BG, fg=C_TEXT,
                 font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(6, 2))
        lf2 = tk.Frame(s4, bg=C_BG)
        lf2.pack(fill=tk.X)
        for lbl, val in [("English","en"),("Arabic+EN","ar"),("French+EN","fr")]:
            tk.Radiobutton(lf2, text=lbl, variable=self.lang_choice, value=val,
                           bg=C_BG, fg=C_TEXT, selectcolor="white",
                           font=("Segoe UI", 8),
                           activebackground=C_BG).pack(side=tk.LEFT, padx=(0, 8))

        tk.Label(s4, text="Decoder:", bg=C_BG, fg=C_TEXT,
                 font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(6, 2))
        df2 = tk.Frame(s4, bg=C_BG)
        df2.pack(fill=tk.X)
        for lbl, val in [("Greedy (fast)","greedy"),("Beam Search (accurate)","beamsearch")]:
            tk.Radiobutton(df2, text=lbl, variable=self.decoder_choice, value=val,
                           bg=C_BG, fg=C_TEXT, selectcolor="white",
                           font=("Segoe UI", 8),
                           activebackground=C_BG).pack(anchor="w")

        tk.Checkbutton(s4, text="Use GPU (CUDA)", variable=self.gpu_enabled,
                       bg=C_BG, fg=C_TEXT, selectcolor="white",
                       font=("Segoe UI", 8),
                       activebackground=C_BG).pack(anchor="w", pady=(6, 0))

        # 06 Preview
        s5 = section("06  Preview", C_PURPLE)
        tk.Checkbutton(s5, text="Show bounding boxes",
                       variable=self.show_boxes, bg=C_BG, fg=C_TEXT,
                       selectcolor="white", font=("Segoe UI", 9),
                       activebackground=C_BG,
                       command=self._refresh_preview).pack(anchor="w")
        tk.Checkbutton(s5, text="Overlay text on preview",
                       variable=self.show_text_ov, bg=C_BG, fg=C_TEXT,
                       selectcolor="white", font=("Segoe UI", 9),
                       activebackground=C_BG,
                       command=self._refresh_preview).pack(anchor="w")

        # 07 Actions
        s6 = section("07  Actions", C_ACCENT)
        self.btn_init = btn(s6, "INITIALIZE OCR ENGINE", C_BLUE, self.init_ocr)
        self.btn_init.pack(fill=tk.X, pady=(0, 5))
        self.btn_single = btn(s6, "TEST CURRENT IMAGE", C_PURPLE, self.process_single, py=6)
        self.btn_single.pack(fill=tk.X, pady=(0, 5))
        self.btn_run = btn(s6, "PROCESS ALL IMAGES", C_ACCENT, self.run_all)
        self.btn_run.pack(fill=tk.X, pady=(0, 5))
        self.btn_run.config(state="disabled")
        self.btn_export = btn(s6, "EXPORT TO EXCEL", C_ORANGE, self.export_excel)
        self.btn_export.pack(fill=tk.X, pady=(0, 5))
        self.btn_export.config(state="disabled")
        self.btn_open_folder = btn(s6, "OPEN OUTPUT FOLDER", "#636e72",
                                    self.open_output_folder, py=5)
        self.btn_open_folder.pack(fill=tk.X)
        self.lbl_hint = tk.Label(s6, text="Initialize OCR engine first",
                                  bg=C_BG, fg=C_GRAY, font=("Segoe UI", 8),
                                  wraplength=280, justify=tk.LEFT)
        self.lbl_hint.pack(anchor="w", pady=(5, 0))

        tk.Frame(left, bg=C_BG, height=20).pack()

    def _nbtn(self, p, t, c):
        return tk.Button(p, text=t, font=("Segoe UI", 9, "bold"),
                          bg="#3d3d6b", fg="white",
                          activebackground="#252545",
                          relief=tk.FLAT, bd=0, padx=12, pady=5,
                          cursor="hand2", command=c)

    def _toggle_ai_ui(self):
        if self.ai_enabled.get():
            self.ai_frame.pack(fill=tk.X)
        else:
            self.ai_frame.pack_forget()

    # ═══════════════════════ ROI ═══════════════════════

    def _canvas_to_img(self, cx, cy):
        ox, oy = self._disp_offset
        ix = (cx - ox) / self._disp_scale
        iy = (cy - oy) / self._disp_scale
        return ix, iy

    def _roi_on_press(self, e):
        self._roi_drawing = True
        self._roi_start   = (e.x, e.y)
        if self._roi_rect_id:
            self.canvas.delete(self._roi_rect_id)

    def _roi_on_drag(self, e):
        if not self._roi_drawing or not self._roi_start:
            return
        if self._roi_rect_id:
            self.canvas.delete(self._roi_rect_id)
        x0, y0 = self._roi_start
        self._roi_rect_id = self.canvas.create_rectangle(
            x0, y0, e.x, e.y, outline="#fdcb6e", width=2, dash=(6, 3))

    def _roi_on_release(self, e):
        if not self._roi_drawing or not self._roi_start:
            return
        self._roi_drawing = False
        x0c, y0c = self._roi_start
        x1c, y1c = e.x, e.y
        x0c, x1c = min(x0c, x1c), max(x0c, x1c)
        y0c, y1c = min(y0c, y1c), max(y0c, y1c)
        if abs(x1c - x0c) < 10 or abs(y1c - y0c) < 10:
            return
        x0i, y0i = self._canvas_to_img(x0c, y0c)
        x1i, y1i = self._canvas_to_img(x1c, y1c)
        if self._current_cv_img is not None:
            h, w = self._current_cv_img.shape[:2]
            x0i, y0i = max(0, int(x0i)), max(0, int(y0i))
            x1i, y1i = min(w, int(x1i)), min(h, int(y1i))
            self._roi = (x0i, y0i, x1i, y1i)
            self.lbl_roi_coords.config(
                text=f"({x0i},{y0i}) -> ({x1i},{y1i})", fg="#fdcb6e")
            self.lbl_roi_info.config(
                text=f"  ROI: ({x0i},{y0i}) -> ({x1i},{y1i})  |  Draw again to change")
            self._refresh_preview()

    def clear_roi(self):
        self._roi = None
        if self._roi_rect_id:
            self.canvas.delete(self._roi_rect_id)
            self._roi_rect_id = None
        self.lbl_roi_coords.config(text="No region set  (full image)", fg=C_GRAY)
        self.lbl_roi_info.config(text="  Drag on image to select data region (ROI)")
        self._refresh_preview()

    # ═══════════════════════ BROWSE ═══════════════════════

    def browse_input(self):
        d = filedialog.askdirectory(title="Select screenshots folder")
        if d:
            self.input_folder.set(d)
            self._load_image_list(d)
            if not self.output_folder.get():
                out = str(Path(d).parent / "ocr_output")
                self.output_folder.set(out)
                self.lbl_out_path.config(text=out)

    def browse_output(self):
        d = filedialog.askdirectory(title="Select output folder")
        if d:
            self.output_folder.set(d)
            self.lbl_out_path.config(text=d)

    def _load_image_list(self, folder):
        exts = {".png", ".jpg", ".jpeg", ".bmp", ".tiff"}
        self.image_files = sorted([
            str(p) for p in Path(folder).iterdir()
            if p.suffix.lower() in exts
        ])
        self.current_idx = 0
        n = len(self.image_files)
        self.lbl_stats.config(text=f"Found {n} image(s).", fg=C_TEXT)
        self._set_status(f"Loaded {n} images")
        if self.image_files:
            self._show_image_raw(self.image_files[0])
            self.lbl_progress_nav.config(text=f"1 / {n}")

    # ═══════════════════════ OCR INIT ═══════════════════════

    def init_ocr(self):
        self.btn_init.config(state="disabled", text="Loading…  please wait")
        self.lbl_hint.config(text="Downloading model on first run (~100 MB)…", fg=C_ORANGE)
        self._set_status("Initializing EasyOCR…")
        threading.Thread(target=self._do_init_ocr, daemon=True).start()

    def _do_init_ocr(self):
        try:
            langs = [self.lang_choice.get()]
            if self.lang_choice.get() in ("ar", "fr"):
                langs.append("en")
            self.reader = easyocr.Reader(langs, gpu=self.gpu_enabled.get(), verbose=False)
            self.root.after(0, self._ocr_init_done)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("OCR Error", str(e)))
            self.root.after(0, lambda: self.btn_init.config(
                state="normal", text="INITIALIZE OCR ENGINE"))

    def _ocr_init_done(self):
        self.btn_init.config(state="normal", text="OCR ENGINE READY")
        self.btn_run.config(state="normal")
        self.lbl_hint.config(
            text="OCR ready!\n1. Draw ROI  2. Test image  3. Process all", fg=C_ACCENT)
        self._set_status("EasyOCR ready")

    # ═══════════════════════ AI CONNECTION ═══════════════════════

    def test_ai_connection(self):
        if not GEMINI_AVAILABLE:
            messagebox.showerror("Missing Library",
                                  "Run:  pip install google-generativeai")
            return
        key = self.gemini_api_key.get().strip()
        if not key:
            messagebox.showwarning("No Key", "Enter your Gemini API key.")
            return
        self.lbl_ai_status.config(text="Testing connection…", fg=C_GOLD)
        threading.Thread(target=self._do_test_ai, args=(key,), daemon=True).start()

    def _do_test_ai(self, key):
        try:
            genai.configure(api_key=key)
            model = genai.GenerativeModel(self.gemini_model.get())
            resp  = model.generate_content("Reply with exactly: OK")
            txt   = resp.text.strip()
            if "OK" in txt:
                self.root.after(0, lambda: self.lbl_ai_status.config(
                    text="Connected! Gemini is ready.", fg=C_ACCENT))
                self._gemini_client = model
            else:
                self.root.after(0, lambda: self.lbl_ai_status.config(
                    text=f"Unexpected reply: {txt}", fg=C_ORANGE))
        except Exception as e:
            self.root.after(0, lambda: self.lbl_ai_status.config(
                text=f"Error: {e}", fg=C_RED))

    def _get_gemini_client(self):
        """Return a fresh Gemini client (re-creates if needed)."""
        if not GEMINI_AVAILABLE:
            return None
        key = self.gemini_api_key.get().strip()
        if not key:
            return None
        genai.configure(api_key=key)
        return genai.GenerativeModel(self.gemini_model.get())

    # ═══════════════════════ AI CORRECTION ═══════════════════════

    def _ai_correct_rows(self, image_np, raw_rows):
        """
        Send the (cropped) image + raw OCR rows to Gemini.
        Returns a list of corrected row strings (same length as raw_rows).
        """
        client = self._get_gemini_client()
        if client is None:
            return None

        # Encode image as JPEG bytes
        ok, buf = cv2.imencode(".jpg", image_np, [cv2.IMWRITE_JPEG_QUALITY, 92])
        if not ok:
            return None
        img_bytes = buf.tobytes()

        # Build numbered raw OCR text
        numbered = "\n".join(
            f"{i+1}. {r['text']}" for i, r in enumerate(raw_rows)
        )

        prompt = f"""You are correcting OCR output from an airline reservation system screenshot.
The data is a list of passenger records. Each row has this fixed structure:
  [ROW_NUMBER]  [PASSENGER_NAME/TITLE]  [TICKET_CODE]  [SPACE_CODE]  [STATUS_HK/HX/etc]  [DATE_DDMMM]  [FLIGHT_CODE]

The OCR made these common errors:
- Extra spaces inside words (e.g. "A L NAHDI" -> "ALNAHDI")
- Confused characters: 0/O, 1/I/l, 5/S, 8/B, Q/0, n/N, etc.
- Row numbers may have extra spaces (e.g. "0 2 0" -> "020")
- The row number is always a 3-digit number (020, 021, 022…)

Here is the raw OCR output ({len(raw_rows)} rows):
{numbered}

Look carefully at the image to correct each row. Return ONLY a JSON array of corrected strings, one per row, in the same order. No extra text, no markdown. Example format:
["020  03ALNAHDI/NOUF MRS  7CTIPR Y  HK  02FEB  MEDSV0421", "021  ..."]"""

        try:
            import PIL.Image as PILImage
            import io
            pil_img = PILImage.open(io.BytesIO(img_bytes))
            response = client.generate_content([prompt, pil_img])
            raw_json = response.text.strip()

            # Strip markdown code fences if present
            raw_json = re.sub(r"^```[a-z]*\n?", "", raw_json)
            raw_json = re.sub(r"\n?```$", "", raw_json.strip())

            corrected = json.loads(raw_json)
            if isinstance(corrected, list) and len(corrected) == len(raw_rows):
                return corrected
            # If lengths differ, pad or trim
            while len(corrected) < len(raw_rows):
                corrected.append(raw_rows[len(corrected)]["text"])
            return corrected[:len(raw_rows)]

        except Exception as e:
            self.root.after(0, lambda: self.lbl_ai_status.config(
                text=f"AI error: {e}", fg=C_RED))
            return None

    # ═══════════════════════ ROW DETECTION ═══════════════════════

    def _get_roi_crop(self, image_path):
        img = cv2.imread(image_path)
        if img is None:
            return None, None, (0, 0)
        orig = img.copy()
        if self._roi is not None:
            x1, y1, x2, y2 = self._roi
            h, w = img.shape[:2]
            x1, y1 = max(0, x1), max(0, y1)
            x2, y2 = min(w, x2), min(h, y2)
            if x2 > x1 and y2 > y1:
                return img[y1:y2, x1:x2], orig, (x1, y1)
        return img, orig, (0, 0)

    def _run_ocr_on_image(self, image_path):
        cropped, orig, (ox, oy) = self._get_roi_crop(image_path)
        if cropped is None:
            return None, [], []

        tmp_path = str(Path(image_path).parent / "__ocr_tmp__.png")
        cv2.imwrite(tmp_path, cropped)

        try:
            result = self.reader.readtext(
                tmp_path, detail=1, paragraph=False, min_size=8,
                text_threshold=self.text_threshold.get(),
                link_threshold=0.3, low_text=0.3,
                width_ths=self.width_threshold.get(),
                height_ths=0.6, ycenter_ths=0.6,
                decoder=self.decoder_choice.get(),
                beamWidth=5, add_margin=0.1,
            )
        except TypeError:
            result = self.reader.readtext(tmp_path, detail=1,
                                           paragraph=False, min_size=8)
        try:
            os.remove(tmp_path)
        except Exception:
            pass

        conf_thr = self.min_confidence.get()
        min_len  = self.min_row_len.get()

        filtered = [(bbox, text, conf) for bbox, text, conf in result
                    if conf >= conf_thr and len(text.strip()) >= min_len]
        if not filtered:
            return orig, [], []

        def yc(item):
            b = item[0]
            return (b[0][1] + b[2][1]) / 2

        filtered.sort(key=yc)
        tol = self.row_tolerance.get()
        rows, cur = [], [filtered[0]]
        for item in filtered[1:]:
            if abs(yc(item) - yc(cur[-1])) <= tol:
                cur.append(item)
            else:
                rows.append(cur)
                cur = [item]
        rows.append(cur)

        row_data = []
        for row_items in rows:
            row_items.sort(key=lambda x: x[0][0][0])
            row_text = "  ".join(item[1] for item in row_items)
            xs = [pt[0] + ox for item in row_items for pt in item[0]]
            ys = [pt[1] + oy for item in row_items for pt in item[0]]
            row_data.append({
                "text":         row_text,
                "text_ai":      None,         # filled in later by AI
                "box":          (min(xs), min(ys), max(xs), max(ys)),
                "confidence":   sum(i[2] for i in row_items) / len(row_items),
                "num_words":    len(row_items),
                "ai_corrected": False,
            })

        # ── AI CORRECTION (optional) ──────────────────────
        if self.ai_enabled.get() and row_data:
            self._set_status("AI correcting rows…")
            corrected_texts = self._ai_correct_rows(cropped, row_data)
            if corrected_texts:
                for i, row in enumerate(row_data):
                    original = row["text"]
                    fixed    = corrected_texts[i] if i < len(corrected_texts) else original
                    row["text_ai"]      = fixed
                    row["ai_corrected"] = (fixed.strip() != original.strip())

        return orig, row_data, filtered

    # ═══════════════════════ PREVIEW ═══════════════════════

    def _show_image_raw(self, path):
        img = cv2.imread(path)
        if img is None:
            return
        self._current_cv_img   = img
        self._current_row_data = []
        self._render_preview(img, [])
        self.lbl_filename.config(text=Path(path).name)

    def _render_preview(self, cv_img, row_data):
        display = cv_img.copy()
        if len(display.shape) == 2:
            display = cv2.cvtColor(display, cv2.COLOR_GRAY2BGR)

        # ROI rect
        if self._roi is not None:
            x1, y1, x2, y2 = self._roi
            cv2.rectangle(display, (x1, y1), (x2, y2), (253, 203, 110), 3)
            cv2.putText(display, "DATA REGION", (x1, max(y1 - 8, 10)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (253, 203, 110), 2)

        if self.show_boxes.get() and row_data:
            # Decide whether to use AI text or raw
            use_ai = self.ai_enabled.get() and self._active_tab.get() == "ai"

            for i, row in enumerate(row_data):
                ai_fixed = row.get("ai_corrected", False) and use_ai
                color    = ROW_AI_COLOR if ai_fixed else ROW_COLORS[i % len(ROW_COLORS)]
                x1, y1, x2, y2 = [int(v) for v in row["box"]]

                ov = display.copy()
                cv2.rectangle(ov, (x1, y1), (x2, y2), color, -1)
                cv2.addWeighted(ov, 0.15, display, 0.85, 0, display)
                cv2.rectangle(display, (x1, y1), (x2, y2), color, 2)

                label = f"#{i+1}" + (" *" if ai_fixed else "")
                bx, by = max(0, x1), max(18, y1)
                cv2.rectangle(display, (bx, by - 16),
                              (bx + len(label)*8 + 6, by), color, -1)
                cv2.putText(display, label, (bx + 3, by - 3),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.4,
                            (0, 0, 0), 1, cv2.LINE_AA)

                if self.show_text_ov.get():
                    txt = (row.get("text_ai") or row["text"]) if use_ai else row["text"]
                    short = txt[:55] + ("..." if len(txt) > 55 else "")
                    cv2.putText(display, short, (x1, y2 + 13),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.35,
                                color, 1, cv2.LINE_AA)

        cw = max(self.canvas.winfo_width(),  600)
        ch = max(self.canvas.winfo_height(), 400)
        h, w = display.shape[:2]
        scale  = min(cw / w, ch / h, 1.0)
        nw, nh = int(w * scale), int(h * scale)
        self._disp_scale  = scale
        self._disp_offset = ((cw - nw) // 2, (ch - nh) // 2)
        display = cv2.resize(display, (nw, nh), interpolation=cv2.INTER_AREA)
        rgb     = cv2.cvtColor(display, cv2.COLOR_BGR2RGB)
        self._tk_img = ImageTk.PhotoImage(Image.fromarray(rgb))
        self.canvas.delete("all")
        ox, oy = self._disp_offset
        self.canvas.create_image(ox, oy, image=self._tk_img, anchor=tk.NW)

    def _refresh_preview(self):
        if self._current_cv_img is not None:
            self._render_preview(self._current_cv_img, self._current_row_data)

    # ═══════════════════════ NAVIGATION ═══════════════════════

    def prev_image(self):
        if self.image_files and self.current_idx > 0:
            self.current_idx -= 1
            self._show_image_raw(self.image_files[self.current_idx])
            self.lbl_progress_nav.config(
                text=f"{self.current_idx+1} / {len(self.image_files)}")

    def next_image(self):
        if self.image_files and self.current_idx < len(self.image_files) - 1:
            self.current_idx += 1
            self._show_image_raw(self.image_files[self.current_idx])
            self.lbl_progress_nav.config(
                text=f"{self.current_idx+1} / {len(self.image_files)}")

    # ═══════════════════════ SINGLE TEST ═══════════════════════

    def process_single(self):
        if not self.reader:
            messagebox.showwarning("OCR", "Initialize OCR engine first.")
            return
        if not self.image_files:
            messagebox.showwarning("No images", "Load a folder first.")
            return
        path = self.image_files[self.current_idx]
        self._set_status(f"Processing {Path(path).name}…")
        threading.Thread(target=self._do_single, args=(path,), daemon=True).start()

    def _do_single(self, path):
        img, row_data, _ = self._run_ocr_on_image(path)
        self.root.after(0, lambda: self._display_result(img, row_data, path))

    def _display_result(self, img, row_data, path):
        self._current_cv_img   = img
        self._current_row_data = row_data
        self._render_preview(img, row_data)
        self._show_text_output(row_data, Path(path).name)
        n_ai = sum(1 for r in row_data if r.get("ai_corrected"))
        lbl  = f"Rows: {len(row_data)}"
        if n_ai:
            lbl += f"  |  AI fixed: {n_ai}"
        self.lbl_row_count.config(text=lbl)
        self._set_status(f"Done  -  {len(row_data)} rows" + (f"  ({n_ai} AI-corrected)" if n_ai else ""))

    def _show_text_output(self, row_data, filename):
        # RAW OCR tab
        self.txt_ocr.config(state="normal")
        self.txt_ocr.delete("1.0", tk.END)
        self.txt_ocr.insert(tk.END, f"FILE: {filename}  [RAW OCR]\n")
        self.txt_ocr.insert(tk.END, "-" * 72 + "\n")
        for i, row in enumerate(row_data, 1):
            pct = f"{row['confidence']*100:.0f}%"
            self.txt_ocr.insert(tk.END, f"ROW {i:>3}  [{pct}]  {row['text']}\n")
        self.txt_ocr.config(state="disabled")

        # AI CORRECTED tab
        self.txt_ai.config(state="normal")
        self.txt_ai.delete("1.0", tk.END)
        has_ai = any(r.get("text_ai") for r in row_data)
        if has_ai:
            self.txt_ai.insert(tk.END, f"FILE: {filename}  [AI CORRECTED]  (* = changed)\n")
            self.txt_ai.insert(tk.END, "-" * 72 + "\n")
            for i, row in enumerate(row_data, 1):
                ai_txt  = row.get("text_ai") or row["text"]
                changed = row.get("ai_corrected", False)
                marker  = " *" if changed else "  "
                pct     = f"{row['confidence']*100:.0f}%"
                self.txt_ai.insert(tk.END,
                    f"ROW {i:>3}  [{pct}]{marker}  {ai_txt}\n")
        else:
            self.txt_ai.insert(tk.END,
                "AI correction not run.\n\nEnable AI Correction, enter API key, and process the image.")
        self.txt_ai.config(state="disabled")

    # ═══════════════════════ PROCESS ALL ═══════════════════════

    def run_all(self):
        if not self.reader:
            messagebox.showwarning("OCR", "Initialize OCR engine first.")
            return
        if not self.image_files:
            messagebox.showwarning("No images", "Load a folder first.")
            return
        if self.processing:
            return
        self.processing  = True
        self.all_results = []
        self.btn_run.config(state="disabled", text="Processing…")
        self.btn_export.config(state="disabled")
        threading.Thread(target=self._do_run_all, daemon=True).start()

    def _do_run_all(self):
        total = len(self.image_files)
        for idx, path in enumerate(self.image_files):
            self.root.after(0, lambda p=path, i=idx:
                            self._set_status(f"Processing {i+1}/{total}  -  {Path(p).name}"))
            self.root.after(0, lambda i=idx:
                            self.progress_bar.configure(value=(i / total) * 100))

            img, row_data, _ = self._run_ocr_on_image(path)
            fname = Path(path).name

            for row_idx, row in enumerate(row_data, 1):
                # Use AI text if available, else raw
                final_text = (row.get("text_ai") or row["text"]) \
                             if (self.ai_enabled.get() and row.get("text_ai")) \
                             else row["text"]
                self.all_results.append({
                    "Source File":    fname,
                    "Row #":          row_idx,
                    "Raw OCR Text":   row["text"],
                    "Final Text":     final_text,
                    "AI Corrected":   "Yes" if row.get("ai_corrected") else "No",
                    "Confidence %":   round(row["confidence"] * 100, 1),
                    "Word Count":     row["num_words"],
                    "Box X1": int(row["box"][0]), "Box Y1": int(row["box"][1]),
                    "Box X2": int(row["box"][2]), "Box Y2": int(row["box"][3]),
                })

            if img is not None:
                self.root.after(0, lambda im=img, rd=row_data, p=path, i=idx:
                                self._update_live(im, rd, p, i, total))

        self.root.after(0, self._all_done)

    def _update_live(self, img, row_data, path, idx, total):
        self._current_cv_img   = img
        self._current_row_data = row_data
        self._render_preview(img, row_data)
        self._show_text_output(row_data, Path(path).name)
        self.current_idx = idx
        self.lbl_filename.config(text=Path(path).name)
        self.lbl_progress_nav.config(text=f"{idx+1} / {total}")
        n_ai = sum(1 for r in row_data if r.get("ai_corrected"))
        self.lbl_row_count.config(
            text=f"Rows: {len(row_data)}  |  Total: {len(self.all_results)}" +
                 (f"  |  AI fixed: {n_ai}" if n_ai else ""))

    def _all_done(self):
        self.processing = False
        self.progress_bar.configure(value=100)
        self.btn_run.config(state="normal", text="PROCESS ALL IMAGES")
        self.btn_export.config(state="normal")
        n = len(self.all_results)
        ai_rows = sum(1 for r in self.all_results if r.get("AI Corrected") == "Yes")
        msg = f"Done!  {n} rows extracted."
        if ai_rows:
            msg += f"\n{ai_rows} rows AI-corrected."
        self.lbl_hint.config(text=msg + "\nClick Export to save.", fg=C_ACCENT)
        messagebox.showinfo("Processing Complete",
                            f"Processed {len(self.image_files)} images.\n"
                            f"Extracted {n} rows.\n"
                            + (f"AI corrected: {ai_rows} rows.\n" if ai_rows else "") +
                            "\nClick  EXPORT TO EXCEL  to save.")

    # ═══════════════════════ EXCEL EXPORT ═══════════════════════

    def export_excel(self):
        if not self.all_results:
            messagebox.showwarning("No data", "Run processing first.")
            return
        out_dir = os.path.abspath(self.output_folder.get().strip()
                                   or str(Path.home() / "Desktop" / "ocr_output"))
        try:
            os.makedirs(out_dir, exist_ok=True)
            # Write test
            tf = os.path.join(out_dir, "_wtest_.tmp")
            with open(tf, "w") as f: f.write("x")
            os.remove(tf)
        except Exception as e:
            messagebox.showerror("Write Error",
                                  f"Cannot write to:\n{out_dir}\n\n{e}")
            return

        out_path = os.path.join(out_dir, "ocr_extracted_rows.xlsx")
        self._set_status(f"Saving…  {out_path}")

        try:
            df = pd.DataFrame(self.all_results)

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="All Rows")
                ws = writer.sheets["All Rows"]

                HDR_FILL  = PatternFill("solid", fgColor="1a1d2e")
                ALT_FILL  = PatternFill("solid", fgColor="252840")
                EVEN_FILL = PatternFill("solid", fgColor="2e3250")
                AI_FILL   = PatternFill("solid", fgColor="4a3f8a")   # highlight AI rows
                HDR_FONT  = Font(bold=True, color="00b894", name="Consolas", size=10)
                DATA_FONT = Font(color="e8eaf6", name="Consolas", size=9)
                AI_FONT   = Font(color="a29bfe", name="Consolas", size=9, bold=True)
                HDR_BRD   = Border(bottom=Side(style="thin", color="00b894"))

                COL_W = {
                    "Source File": 28, "Row #": 6,
                    "Raw OCR Text": 70, "Final Text": 70,
                    "AI Corrected": 12, "Confidence %": 12,
                    "Word Count": 10,
                    "Box X1": 7, "Box Y1": 7, "Box X2": 7, "Box Y2": 7,
                }

                for ci, cn in enumerate(df.columns, 1):
                    c = ws.cell(row=1, column=ci)
                    c.fill      = HDR_FILL
                    c.font      = HDR_FONT
                    c.border    = HDR_BRD
                    c.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(cn, 10)

                ai_col_idx = list(df.columns).index("AI Corrected") + 1

                for ri in range(2, len(df) + 2):
                    is_ai = ws.cell(row=ri, column=ai_col_idx).value == "Yes"
                    fill  = AI_FILL if (is_ai and self._show_ai_diff.get()) else \
                            (EVEN_FILL if ri % 2 == 0 else ALT_FILL)
                    font  = AI_FONT if is_ai else DATA_FONT
                    for ci in range(1, len(df.columns) + 1):
                        c = ws.cell(row=ri, column=ci)
                        c.fill      = fill
                        c.font      = font
                        c.alignment = Alignment(horizontal="left",
                                                 wrap_text=(ci in (3, 4)))

                ws.freeze_panes    = "A2"
                ws.auto_filter.ref = ws.dimensions

                # Summary
                summ = (df.groupby("Source File")
                          .agg(Total_Rows    =("Row #",        "count"),
                               Avg_Confidence=("Confidence %", "mean"),
                               AI_Corrected  =("AI Corrected", lambda x: (x=="Yes").sum()))
                          .reset_index())
                summ["Avg_Confidence"] = summ["Avg_Confidence"].round(1)
                summ.to_excel(writer, index=False, sheet_name="Summary")

        except PermissionError:
            messagebox.showerror("Permission Error",
                                  f"File open in Excel:\n{out_path}\nClose it first.")
            return
        except Exception as e:
            messagebox.showerror("Export Error", f"{e}")
            return

        if not os.path.exists(out_path):
            messagebox.showerror("Error", f"File not created:\n{out_path}")
            return

        sz = os.path.getsize(out_path)
        self._last_excel_path = out_path
        self._set_status(f"Saved  ({sz:,} bytes)  ->  {out_path}")
        self.lbl_out_path.config(text=out_path)
        self.lbl_hint.config(text=f"Saved: {os.path.basename(out_path)} ({sz:,} bytes)",
                              fg=C_ACCENT)
        messagebox.showinfo("Export Complete",
                            f"Excel saved!\n\n{out_path}\n\n"
                            f"Size: {sz:,} bytes  |  Rows: {len(self.all_results)}\n\n"
                            "Click  OPEN OUTPUT FOLDER  to view.")

    def open_output_folder(self):
        out_dir = self.output_folder.get().strip()
        if not out_dir or not os.path.isdir(out_dir):
            if self._last_excel_path:
                out_dir = str(Path(self._last_excel_path).parent)
            else:
                messagebox.showinfo("Open Folder", "Export Excel first.")
                return
        try:
            if sys.platform == "win32":   os.startfile(out_dir)
            elif sys.platform == "darwin": subprocess.Popen(["open", out_dir])
            else:                          subprocess.Popen(["xdg-open", out_dir])
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ═══════════════════════ UTILS ═══════════════════════

    def _set_status(self, msg):
        self.root.after(0, lambda: self.lbl_status.config(text=msg))


if __name__ == "__main__":
    root = tk.Tk()
    app  = Phase2OCRExtractor(root)
    root.mainloop()